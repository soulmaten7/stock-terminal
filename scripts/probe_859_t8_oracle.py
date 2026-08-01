"""STEP 859 §3 — T8 원본을 정답지(oracle)로. openpyxl 캐시값으로 도미노를 앵커, PIE 수식을 파이썬 재구현해 3케이스 정답 생성."""
import openpyxl, json

WB = "data/sources/expectations-investing/T8.xlsx"

# ── T8 캐시값(도미노) 읽기 — data_only=True는 엑셀이 마지막 저장한 계산값 ──
wb = openpyxl.load_workbook(WB, data_only=True)
ws = wb["Price Implied Expectations"]
c31 = ws["C31"].value  # MIFP
c33 = ws["C33"].value  # 25년가치/주가 (25+일 때만)
# D27:AB27 = 주당가치 1~25년
cached_ps = [ws.cell(row=27, column=col).value for col in range(4, 29)]  # D=4 .. AB=28
print("=== T8 캐시값(도미노) ===")
print(f"  C31(MIFP)={c31}  C33={c33}")
print(f"  D27:AB27 주당가치 처음3={[round(x,2) if isinstance(x,(int,float)) else x for x in cached_ps[:3]]} ... 25년={round(cached_ps[-1],2) if isinstance(cached_ps[-1],(int,float)) else cached_ps[-1]}")

# ── PIE 수식 재구현 (857에서 셀 판독한 그대로) ──
def pie(startingSales, g, opMargin, startMargin, tax, fixedRate, workRate, wacc, inflation, price, shares, debt, nonOp, maxYears=25):
    prevSales = startingSales; cumPv = 0.0; ps = []
    for t in range(1, maxYears + 1):
        sales = prevSales * (1 + g)
        op = sales * opMargin
        nopat = op * (1 - tax)
        incF = (sales - prevSales) * fixedRate
        incW = (sales - prevSales) * workRate
        fcf = nopat - incF - incW
        disc = 1 / (1 + wacc) ** t
        cumPv += fcf * disc
        pvRes = (nopat * (1 + inflation)) / (wacc - inflation) * disc
        corp = cumPv + pvRes
        sh = corp + nonOp - debt
        ps.append(sh / shares)
        prevSales = sales
    # 단조성 (증분ROIC<WACC면 down=value_destroying)
    down = all(ps[k] <= ps[k-1] + 1e-9 for k in range(1, len(ps)))
    up = all(ps[k] >= ps[k-1] - 1e-9 for k in range(1, len(ps)))
    if down and not up:
        return {"verdict": "value_destroying", "gap": None, "explained": None, "ps25": ps[-1]}
    # T8 C31: LOOKUP(price, D27:AB27, D4:AB4)
    if price > ps[-1]:
        return {"verdict": "25+", "gap": None, "explained": ps[-1] / price, "ps25": ps[-1]}
    if price < ps[0]:
        return {"verdict": "<1", "gap": None, "explained": None, "ps25": ps[-1]}
    gap = 1
    for k in range(len(ps)):
        if ps[k] <= price:
            gap = k + 1
    return {"verdict": "years", "gap": gap, "explained": None, "ps25": ps[-1]}

DOM = dict(startingSales=3618.8, g=0.07, opMargin=0.175, startMargin=0.1739, tax=0.165,
           fixedRate=0.15, workRate=0.10, wacc=0.05357, inflation=0.016, price=418,
           shares=39.35, debt=4170, nonOp=391.9)

# 앵커: 재구현이 캐시 도미노와 맞나 (per-share 배열 + MIFP)
r = pie(**DOM)
reimpl_ps = [ (lambda ss=DOM: None)() ]  # placeholder
# per-share 배열 다시 뽑기(대조용)
def pie_ps(**k):
    ps=[]; prevSales=k['startingSales']; cumPv=0.0
    for t in range(1, k.get('maxYears',25)+1):
        sales=prevSales*(1+k['g']); op=sales*k['opMargin']; nopat=op*(1-k['tax'])
        incF=(sales-prevSales)*k['fixedRate']; incW=(sales-prevSales)*k['workRate']
        fcf=nopat-incF-incW; disc=1/(1+k['wacc'])**t; cumPv+=fcf*disc
        pvRes=(nopat*(1+k['inflation']))/(k['wacc']-k['inflation'])*disc
        ps.append((cumPv+pvRes+k['nonOp']-k['debt'])/k['shares']); prevSales=sales
    return ps
rp = pie_ps(**DOM)
maxdiff = max(abs(rp[i]-cached_ps[i]) for i in range(25) if isinstance(cached_ps[i],(int,float)))
print(f"\n=== 앵커: 재구현 vs 캐시 도미노 per-share 최대오차 = {maxdiff:.4f} (0에 가까워야) · 재구현 MIFP={r['gap']} (캐시 {c31}) ===")

# ── 3 케이스 정답 생성 ──
cases = {
    "정상(도미노 8년)": DOM,
    "25+ 유발(WACC 0.15)": {**DOM, "wacc": 0.15},
    "<1 유발(주가 50)": {**DOM, "price": 50},
}
out = {}
print("\n=== 3케이스 정답(파이썬 재구현) ===")
for name, k in cases.items():
    res = pie(**k)
    out[name] = res
    ex = f"{res['explained']*100:.2f}%" if res['explained'] is not None else "—"
    print(f"  {name}: verdict={res['verdict']} gap={res['gap']} 25년가치/주가={ex} (25년가치=${res['ps25']:.2f})")
json.dump(out, open("scripts/probe_859_oracle.json","w"), indent=1, default=str)
print("\n(정답 → scripts/probe_859_oracle.json)")
